"""
Sheets client with 3-tier fallback:

  Tier 1 — Google Sheets API (production, when credentials are configured)
  Tier 2 — Local Excel file  (WiMLDS_Master_Sheet.xlsx in project root)
  Tier 3 — Stub             (no-op, for dry-run / CI use)

The Excel fallback lets you run the full pipeline right now without needing
Google Sheets API credentials. The agent reads event data from your .xlsx
and writes results back to the same file.

Column name → field name mapping is derived from Row 2 of the sheet
(the actual header row), so it stays in sync with your Excel automatically.
"""
from __future__ import annotations

import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from wimlds.core.logger import get_logger

logger = get_logger("sheets_client")

# ── Where to look for the Excel file ─────────────────────────────────────────
# Checks these locations in order:
XLSX_SEARCH_PATHS = [
    Path("WiMLDS_Master_Sheet.xlsx"),                      # project root
    Path("wimlds/WiMLDS_Master_Sheet.xlsx"),               # package folder
    Path("../WiMLDS_Master_Sheet.xlsx"),                   # one level up
    Path("data/WiMLDS_Master_Sheet.xlsx"),                 # data/ subfolder
]

# ── Header normalisation ──────────────────────────────────────────────────────
# These renames map the sheet's column names to the field names the agents use.
# Format:  "sheet column name (lowercased)" → "agent field name"
HEADER_REMAP = {
    "row_id":                    "_row_id",          # internal only
    "event_status":              "event_status",
    "highest_qualification":     "speaker_highest_qualification",
    "tier_1_institution":        "speaker_tier1_institution",
    "special_achievements":      "speaker_special_achievements",
    "linkedin_y_n":              "promote_linkedin",
    "facebook_y_n":              "promote_facebook",
    "x_twitter_y_n":             "promote_x",
    "instagram_y_n":             "promote_instagram",
    "wa_groups_y_n":             "promote_wa_groups",
    "individual_wa_y_n":         "promote_wa_individual",
    "ppt___slides_link":         "ppt_link",          # "PPT / Slides Link" normalises oddly
    "announce_sent":             "announce_sent",
    "t_2d_sent":                 "tminus2_sent",
    "t_1d_sent":                 "tminus1_sent",
    "t_2h_sent":                 "tminus2h_sent",
    "post_event_sent":           "post_event_update_sent",
    "wa_groups_posted":          "whatsapp_groups_posted",
    "indiv_wa_posted":           "whatsapp_individual_posted",
    "partners_notified":         "partners_notified",
    "twitter_tweet_id":          "_twitter_tweet_id",
    "x_tweet_id":                "_twitter_tweet_id",
    "tweet_id":                  "_twitter_tweet_id",
    "twitter_post_url":          "_twitter_post_url",
    "x_post_url":                "_twitter_post_url",
    "tweet_url":                 "_twitter_post_url",
    "x_post_status":             "x_post_status",
    "twitter_post_status":       "x_post_status",
    "x_post_text":               "x_post_text",
    "twitter_post_text":         "x_post_text",
    "x_posted_at":               "x_posted_at",
    "twitter_posted_at":         "x_posted_at",
    "x_error":                   "x_error",
    "twitter_error":             "x_error",
}

FIELD_HEADER_LABELS = {
    "_twitter_tweet_id": "Twitter Tweet ID",
    "_twitter_post_url": "Twitter Post URL",
    "x_post_status": "X Post Status",
    "x_post_text": "X Post Text",
    "x_posted_at": "X Posted At",
    "x_error": "X Error",
    "link": "Link",
}

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _normalise(header: str) -> str:
    """Convert a sheet column header to a Python field name."""
    key = header.strip().lower()
    key = re.sub(r"[^a-z0-9]+", "_", key)   # non-alphanum → _
    key = key.strip("_")
    return HEADER_REMAP.get(key, key)


def _col_letter_to_index(col: str) -> int:
    """Convert column letter(s) like 'A', 'BE' to 0-based index."""
    idx = 0
    for ch in col.upper():
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1


def _index_to_col_letter(index: int) -> str:
    index += 1
    result = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        result = chr(65 + rem) + result
    return result


def _parse_cell_ref(ref: str):
    """Split 'BE3' into ('BE', 3)."""
    m = re.match(r"([A-Z]+)(\d+)", ref)
    if not m:
        return None, None
    return m.group(1), int(m.group(2))


# ════════════════════════════════════════════════════════════════════════════
class ExcelSheetsClient:
    """
    Read/write events from a local WiMLDS_Master_Sheet.xlsx file.

    Row layout:
      Row 1  — title banner (merged cell, ignored)
      Row 2  — column headers
      Row 3+ — event data (Row ID in column A)
    """

    def __init__(self, xlsx_path: Path):
        self.xlsx_path = xlsx_path
        self._headers: Optional[dict] = None   # col_letter → field_name
        self._col_map: Optional[dict] = None   # field_name → col_letter
        self._shared_strings: Optional[list[str]] = None
        logger.info(f"ExcelSheetsClient: using {xlsx_path}")

    # ── Internal helpers ─────────────────────────────────────────────────────

    def _read_sheet_xml(self) -> ET.Element:
        with zipfile.ZipFile(self.xlsx_path) as z:
            return ET.fromstring(z.read("xl/worksheets/sheet1.xml"))

    def _load_shared_strings(self) -> list[str]:
        if self._shared_strings is not None:
            return self._shared_strings

        strings: list[str] = []
        try:
            with zipfile.ZipFile(self.xlsx_path) as z:
                if "xl/sharedStrings.xml" not in z.namelist():
                    self._shared_strings = []
                    return self._shared_strings
                root = ET.fromstring(z.read("xl/sharedStrings.xml"))
        except Exception:
            self._shared_strings = []
            return self._shared_strings

        ns = NS
        for si in root.findall(f"{{{ns}}}si"):
            parts = []
            for t_el in si.findall(f".//{{{ns}}}t"):
                parts.append(t_el.text or "")
            strings.append("".join(parts))

        self._shared_strings = strings
        return self._shared_strings

    def _cell_value(self, cell_el: ET.Element) -> str:
        """Extract the string value from a cell element (handles inlineStr and shared strings)."""
        ns = NS
        cell_type = cell_el.get("t", "")
        is_el = cell_el.find(f"{{{ns}}}is")
        if is_el is not None:
            t_el = is_el.find(f"{{{ns}}}t")
            return (t_el.text or "") if t_el is not None else ""
        v_el = cell_el.find(f"{{{ns}}}v")
        raw_value = (v_el.text or "") if v_el is not None else ""
        if cell_type == "s" and raw_value != "":
            try:
                idx = int(raw_value)
                shared = self._load_shared_strings()
                return shared[idx] if 0 <= idx < len(shared) else raw_value
            except Exception:
                return raw_value
        return raw_value

    def _load_headers(self) -> tuple[dict, dict]:
        if self._headers is not None:
            return self._headers, self._col_map

        root = self._read_sheet_xml()
        ns   = NS
        headers  = {}   # col_letter → field_name
        col_map  = {}   # field_name → col_letter

        for row in root.findall(f".//{{{ns}}}row"):
            rnum = int(row.get("r", 0))
            if rnum != 2:
                continue
            for cell in row.findall(f"{{{ns}}}c"):
                ref = cell.get("r", "")
                col, _ = _parse_cell_ref(ref)
                if not col:
                    continue
                raw_header = self._cell_value(cell)
                if not raw_header.strip():
                    continue
                field = _normalise(raw_header)
                headers[col]   = field
                col_map[field] = col
            break

        self._headers = headers
        self._col_map = col_map
        logger.debug(f"Loaded {len(headers)} column headers from Excel")
        return headers, col_map

    def _add_header_column(self, field_name: str) -> str:
        wb = load_workbook(self.xlsx_path)
        ws = wb[wb.sheetnames[0]]

        max_col = ws.max_column
        new_col_idx = max_col + 1
        ws.cell(
            row=2,
            column=new_col_idx,
            value=FIELD_HEADER_LABELS.get(field_name, field_name.strip("_").replace("_", " ").title()),
        )
        wb.save(self.xlsx_path)

        self._headers = None
        self._col_map = None
        self._shared_strings = None
        new_col = _index_to_col_letter(new_col_idx - 1)
        logger.info(f"ExcelSheetsClient: added missing column for field '{field_name}' at {new_col}")
        return new_col

    def _read_row(self, row_number: int) -> dict:
        """Read a single data row. row_number is the Excel row number (3 = first data row)."""
        headers, _ = self._load_headers()
        root = self._read_sheet_xml()
        ns   = NS

        data: dict = {"_row_number": row_number}
        for row in root.findall(f".//{{{ns}}}row"):
            if int(row.get("r", 0)) != row_number:
                continue
            for cell in row.findall(f"{{{ns}}}c"):
                ref = cell.get("r", "")
                col, _ = _parse_cell_ref(ref)
                if col and col in headers:
                    data[headers[col]] = self._cell_value(cell)
            break

        # Fill in any missing fields with empty string
        for field in headers.values():
            data.setdefault(field, "")

        return data

    def _write_cell(self, row_number: int, col_letter: str, value: str):
        """Write a single cell value back into the xlsx file safely."""
        wb = load_workbook(self.xlsx_path)
        ws = wb[wb.sheetnames[0]]
        ws[f"{col_letter}{row_number}"] = str(value)
        wb.save(self.xlsx_path)

    # ── Public API (mirrors GoogleSheetsClient) ───────────────────────────────

    def get_event(self, row_number: int) -> dict:
        """
        Read one event by Excel row number.
        Row 3 = first event (row 2 is headers, row 1 is title).
        """
        data = self._read_row(row_number)
        title = data.get("event_title", data.get("d", "?"))
        logger.info(f"Loaded event from Excel row {row_number}: {title}")
        return data

    def get_all_upcoming(self) -> list[dict]:
        """Return all rows where event_status = 'Upcoming'."""
        headers, _ = self._load_headers()
        root = self._read_sheet_xml()
        ns   = NS

        events = []
        for row in root.findall(f".//{{{ns}}}row"):
            rnum = int(row.get("r", 0))
            if rnum <= 2:
                continue
            data: dict = {"_row_number": rnum}
            for cell in row.findall(f"{{{ns}}}c"):
                ref = cell.get("r", "")
                col, _ = _parse_cell_ref(ref)
                if col and col in headers:
                    data[headers[col]] = self._cell_value(cell)
            for field in headers.values():
                data.setdefault(field, "")
            if data.get("event_status", "").strip().lower() == "upcoming":
                events.append(data)

            logger.info(f"Found {len(events)} upcoming events in Excel")
        return events

    def write_field(self, row_number: int, field_name: str, value: Any) -> bool:
        """Write a single field back to the Excel file."""
        _, col_map = self._load_headers()
        col = col_map.get(field_name)
        if not col:
            try:
                col = self._add_header_column(field_name)
            except Exception as e:
                logger.warning(f"No column found for field '{field_name}' - skipping write ({e})")
                return False
        try:
            self._write_cell(row_number, col, str(value))
            logger.debug(f"Excel write: row={row_number} {field_name}={value} -> col {col}")
            return True
        except Exception as e:
            logger.error(f"Excel write failed for {field_name}: {e}")
            return False

    def write_fields(self, row_number: int, fields: dict) -> bool:
        success = True
        for field_name, value in fields.items():
            if not self.write_field(row_number, field_name, value):
                success = False
        return success

    def set_flag(self, row_number: int, flag: str, value: str = "Y") -> bool:
        return self.write_field(row_number, flag, value)


# ════════════════════════════════════════════════════════════════════════════
class _StubSheetsClient:
    """No-op fallback when neither Google Sheets nor Excel is available."""
    def get_event(self, row_number: int) -> dict:
        logger.warning(f"[STUB] get_event({row_number}) - no sheet configured")
        return {"_row_number": row_number}
    def get_all_upcoming(self) -> list:
        return []
    def write_field(self, *a) -> bool:
        return True
    def write_fields(self, *a) -> bool:
        return True
    def set_flag(self, *a) -> bool:
        return True


# ════════════════════════════════════════════════════════════════════════════
def _build_client():
    """
    Build the best available sheets client:
      1. Google Sheets API (if credentials configured)
      2. Local Excel file   (if WiMLDS_Master_Sheet.xlsx found)
      3. Stub               (fallback)
    """
    # ── Tier 1: Google Sheets API ─────────────────────────────────────────
    try:
        from wimlds.config.settings import settings
        if settings.google_sheets_id and settings.google_service_account_json:
            sa_path = Path(settings.google_service_account_json)
            if sa_path.exists():
                # Import lazily so missing google packages don't break the fallback
                from google.oauth2 import service_account
                from googleapiclient.discovery import build

                SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
                creds   = service_account.Credentials.from_service_account_file(
                    str(sa_path), scopes=SCOPES
                )
                service = build("sheets", "v4", credentials=creds, cache_discovery=False)

                class _GoogleSheetsClient:
                    """Thin wrapper that re-uses the existing SheetsClient logic."""
                    def __init__(self, svc, sheet_id):
                        self._svc      = svc
                        self.sheet_id  = sheet_id

                    def _get_rows(self):
                        resp = self._svc.spreadsheets().values().get(
                            spreadsheetId=self.sheet_id, range="Sheet1!A:BE"
                        ).execute()
                        return resp.get("values", [])

                    def get_event(self, row_number: int) -> dict:
                        rows = self._get_rows()
                        if len(rows) < 2:
                            raise ValueError("Sheet has no header row")
                        headers = [_normalise(h) for h in rows[1]]
                        if row_number - 1 >= len(rows):
                            raise IndexError(f"Row {row_number} not in sheet")
                        vals  = rows[row_number - 1]
                        data  = {headers[i]: (vals[i] if i < len(vals) else "")
                                 for i in range(len(headers))}
                        data["_row_number"] = row_number
                        logger.info(f"Loaded event from Google Sheets row {row_number}: {data.get('event_title','?')}")
                        return data

                    def get_all_upcoming(self) -> list:
                        rows = self._get_rows()
                        if len(rows) < 2:
                            return []
                        headers = [_normalise(h) for h in rows[1]]
                        events  = []
                        for i, row in enumerate(rows[2:], start=3):
                            data = {headers[j]: (row[j] if j < len(row) else "")
                                    for j in range(len(headers))}
                            data["_row_number"] = i
                            if data.get("event_status", "").strip().lower() == "upcoming":
                                events.append(data)
                        return events

                    def write_field(self, row_number: int, field_name: str, value) -> bool:
                        # Build reverse header map on demand
                        rows = self._get_rows()
                        if len(rows) < 2:
                            return False
                        raw_headers = rows[1]
                        for i, h in enumerate(raw_headers):
                            if _normalise(h) == field_name:
                                import string
                                def idx_to_col(n):
                                    col = ""
                                    while n >= 0:
                                        col = chr(n % 26 + ord('A')) + col
                                        n = n // 26 - 1
                                    return col
                                col = idx_to_col(i)
                                cell = f"Sheet1!{col}{row_number}"
                                self._svc.spreadsheets().values().update(
                                    spreadsheetId=self.sheet_id, range=cell,
                                    valueInputOption="USER_ENTERED",
                                    body={"values": [[str(value)]]}
                                ).execute()
                                return True
                        return False

                    def write_fields(self, row_number: int, fields: dict) -> bool:
                        return all(self.write_field(row_number, k, v) for k, v in fields.items())

                    def set_flag(self, row_number: int, flag: str, value="Y") -> bool:
                        return self.write_field(row_number, flag, value)

                client = _GoogleSheetsClient(service, settings.google_sheets_id)
                logger.info("SheetsClient: using Google Sheets API")
                return client
    except Exception as e:
        logger.warning(f"Google Sheets API unavailable ({e}) - trying Excel fallback")

    # ── Tier 2: Local Excel file ──────────────────────────────────────────
    for xlsx_path in XLSX_SEARCH_PATHS:
        if xlsx_path.exists():
            logger.info(f"SheetsClient: using local Excel file -> {xlsx_path}")
            return ExcelSheetsClient(xlsx_path)

    # ── Tier 3: Stub ──────────────────────────────────────────────────────
    logger.warning(
        "SheetsClient: no Google Sheets credentials and no Excel file found.\n"
        "  -> To use Excel: place WiMLDS_Master_Sheet.xlsx in the project root or wimlds/ folder\n"
        "  -> To use Google Sheets: set GOOGLE_SHEETS_ID + GOOGLE_SERVICE_ACCOUNT_JSON in config/.env"
    )
    return _StubSheetsClient()


# Singleton — imported by all agents
sheets_client = _build_client()
